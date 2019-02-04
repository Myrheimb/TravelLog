#Imports all the needed modules to make this program work.
import openpyxl
from openpyxl import workbook
from openpyxl.styles import Font
from time import strftime
from tkinter import *
from config import config

#Creates variables containing the current date, current weekday and current month.
date = strftime('%d.%m.%Y')
weekday = strftime('%A')
month = strftime ('%B')

#Loads in the Excel workbook and selects this month's sheet.
#If this month's sheet doesn't exist, it will create it and write in the headers of each column.
wb = openpyxl.load_workbook(config['excelWorkBook'])
if month in wb.sheetnames:
	ws = wb[month]
else:
	wb.create_sheet(month)
	ws = wb[month]
	ws['A1'] = 'Weekday'
	ws['A1'].font = Font(bold=True)
	ws['B1'] = 'Date'
	ws['B1'].font = Font(bold=True)
	ws['C1'] = 'Workplace'
	ws['C1'].font = Font(bold=True)
	ws['D1'] = 'Travel route'
	ws['D1'].font = Font(bold=True)
	ws['E1'] = 'Distance'
	ws['E1'].font = Font(bold=True)
	ws['F1'] = 'Comment'
	ws['F1'].font = Font(bold=True)

#Defining the GUI.
root = Tk()
root.title('Travel Log')
carIcon = PhotoImage(file='car.png')
root.tk.call('wm', 'iconphoto', root._w, carIcon)

#Creates variables to be filled in by user input and button presses.
travelRoute = ''
totalDist = 0
travelSteps = []
workPlace = ''
textInput = ''

#Defining functions for the different buttons in the program.

#This button is for adding the home address to the variable travelRoute.
def homeButton():
    global travelSteps
    global travelRoute
    travelSteps = travelSteps + [0]
    if travelRoute == '':
        travelRoute = travelRoute + config['Home']
    else:
        travelRoute = travelRoute + ' - ' + config['Home']

#This button is for adding the work1 address to the variable travelRoute.
def work1Button():
    global travelSteps
    global travelRoute
    travelSteps = travelSteps + [1]
    if travelRoute == '':
        travelRoute = travelRoute + config['Work1']
    else:
        travelRoute = travelRoute + ' - ' + config['Work1']

#This button is for adding the work2 address to the variable travelRoute.
def work2Button():
    global travelSteps
    global travelRoute
    travelSteps = travelSteps + [2]
    if travelRoute == '':
        travelRoute = travelRoute + config['Work2']
    else:
        travelRoute = travelRoute + ' - ' + config['Work2']

#This button is for adding the work3 address to the variable travelRoute.
def work3Button():
    global travelSteps
    global travelRoute
    travelSteps = travelSteps + [3]
    if travelRoute == '':
        travelRoute = travelRoute + config['Work3']
    else:
        travelRoute = travelRoute + ' - ' + config['Work3']

#This button is for adding the work4 address to the variable travelRoute.
def work4Button():
    global travelSteps
    global travelRoute
    travelSteps = travelSteps + [4]
    if travelRoute == '':
        travelRoute = travelRoute + config['Work4']
    else:
        travelRoute = travelRoute + ' - ' + config['Work4']

#This button is for adding the work5 address to the variable travelRoute.
def work5Button():
    global travelSteps
    global travelRoute
    travelSteps = travelSteps + [5]
    if travelRoute == '':
        travelRoute = travelRoute + config['Work5']
    else:
        travelRoute = travelRoute + ' - ' + config['Work5']

#This function determines the total travel distance, and which workplace(s) you've worked at.
#Then it appends it to the next row in this month's sheet in your Excel workbook.
def writeToExcel():
	global travelDist
	global travelSteps
	global travelRoute
	global travelSteps
	global textInput
	global workPlace
	global ws
	
	travelSteps.sort()
	
	#These statements determine the total travel distance by looking at which locations you've visited.
	if travelSteps == [0, 0, 1]:
		travelDist = config['distHomeWork1'] + config['distHomeWork1']
	if travelSteps == [0, 0, 2]:
		travelDist = config['distHomeWork2'] + config['distHomeWork2']
	if travelSteps == [0, 0, 3]:
		travelDist = config['distHomeWork3'] + config['distHomeWork3']
	if travelSteps == [0, 0, 4]:
		travelDist = config['distHomeWork4'] + config['distHomeWork4']
	if travelSteps == [0, 0, 5]:
		travelDist = config['distHomeWork5'] + config['distHomeWork5']
	if travelSteps == [0, 0, 1, 2]:
		travelDist = config['distHomeWork1'] + config['distWork1Work2'] + config['distHomeWork2']
	if travelSteps == [0, 0, 1, 3]:
		travelDist = config['distHomeWork1'] + config['distWork1Work3'] + config['distHomeWork3']
	if travelSteps == [0, 0, 1, 4]:
		travelDist = config['distHomeWork1'] + config['distWork1Work4'] + config['distHomeWork4']
	if travelSteps == [0, 0, 1, 5]:
		travelDist = config['distHomeWork1'] + config['distWork1Work5'] + config['distHomeWork5']
	if travelSteps == [0, 0, 2, 3]:
		travelDist = config['distHomeWork2'] + config['distWork2Work3'] + config['distHomeWork3']
	if travelSteps == [0, 0, 2, 4]:
		travelDist = config['distHomeWork2'] + config['distWork2Work4'] + config['distHomeWork4']
	if travelSteps == [0, 0, 2, 5]:
		travelDist = config['distHomeWork2'] + config['distWork2Work5'] + config['distHomeWork5']
	if travelSteps == [0, 0, 3, 4]:
		travelDist = config['distHomeWork3'] + config['distWork3Work4'] + config['distHomeWork4']
	if travelSteps == [0, 0, 3, 5]:
		travelDist = config['distHomeWork3'] + config['distWork3Work5'] + config['distHomeWork5']
	if travelSteps == [0, 0, 4, 5]:
		travelDist = config['distHomeWork4'] + config['distWork4Work5'] + config['distHomeWork5']
	else:
		pass
	
	#These statements determines which workplace you worked at by calculating the total distance traveled.
	if travelDist == config['distHomeWork1'] + config['distHomeWork1']:
		workPlace = config['wpWork1']
	if travelDist == config['distHomeWork2'] + config['distHomeWork2']:
		workPlace = config['wpWork2']
	if travelDist == config['distHomeWork3'] + config['distHomeWork3']:
		workPlace = config['wpWork3']
	if travelDist == config['distHomeWork4'] + config['distHomeWork4']:
		workPlace = config['wpWork4']
	if travelDist == config['distHomeWork5'] + config['distHomeWork5']:
		workPlace = config['wpWork5']
	if travelDist == config['distHomeWork1'] + config['distWork1Work2'] + config['distHomeWork2']:
		workPlace = config['wpWork1'] + ' & ' + config['wpWork2']
	if travelDist == config['distHomeWork1'] + config['distWork1Work3'] + config['distHomeWork3']:
		workPlace = config['wpWork1'] + ' & ' + config['wpWork3']
	if travelDist == config['distHomeWork1'] + config['distWork1Work4'] + config['distHomeWork4']:
		workPlace = config['wpWork1'] + ' & ' + config['wpWork4']
	if travelDist == config['distHomeWork1'] + config['distWork1Work5'] + config['distHomeWork5']:
		workPlace = config['wpWork1'] + ' & ' + config['wpWork5']
	if travelDist == config['distHomeWork2'] + config['distWork2Work3'] + config['distHomeWork3']:
		workPlace = config['wpWork2'] + ' & ' + config['wpWork3']
	if travelDist == config['distHomeWork2'] + config['distWork2Work4'] + config['distHomeWork4']:
		workPlace = config['wpWork2'] + ' & ' + config['wpWork4']
	if travelDist == config['distHomeWork2'] + config['distWork2Work5'] + config['distHomeWork5']:
		workPlace = config['wpWork2'] + ' & ' + config['wpWork5']
	if travelDist == config['distHomeWork3'] + config['distWork3Work4'] + config['distHomeWork4']:
		workPlace = config['wpWork3'] + ' & ' + config['wpWork4']
	if travelDist == config['distHomeWork3'] + config['distWork3Work5'] + config['distHomeWork5']:
		workPlace = config['wpWork3'] + ' & ' + config['wpWork5']
	if travelDist == config['distHomeWork4'] + config['distWork4Work5'] + config['distHomeWork5']:
		workPlace = config['wpWork4'] + ' & ' + config['wpWork5']
	else:
		pass
	
	# Gets the content of the comment field and adds it to the variable textInput.
	textInput = textInputButton.get()
	
	#This appends the collected data into the next row in this month's sheet in your Excel workbook.
	ws.append([weekday, date, workPlace, travelRoute, travelDist, textInput])
	
	#Empty all variables before a possible new append.
	travelRoute = ''
	totalDist = 0
	travelSteps = []
	textInput = ''
	workPlace = ''
	
	#Writes all the data to the Excel workbook.
	wb.save(config['excelWorkBook'])
	
	# Empties the comment field.
	textInputButton.delete(0, 'end')
	
#This function does the same as the above, but also closes the app window.
def writeToExcelClose():
	writeToExcel()
	root.destroy()

#Buttons in the GUI with button titles, sizes and grid definitions.
#Change the text on the buttons b1, b2, b3, b4, b5 and b6 to your own values here.
b1 = Button(root, text='Home', command=homeButton, width=13)
b1.grid(row=0, column=0)
b2 = Button(root, text=config['wpWork1'], command=work1Button, width=13)
b2.grid(row=0, column=1)
b3 = Button(root, text=config['wpWork2'], command=work2Button, width=13)
b3.grid(row=0, column=2)
b4 = Button(root, text=config['wpWork3'], command=work3Button, width=13)
b4.grid(row=1, column=0)
b5 = Button(root, text=config['wpWork4'], command=work4Button, width=13)
b5.grid(row=1, column=1)
b6 = Button(root, text=config['wpWork5'], command=work5Button, width=13)
b6.grid(row=1, column=2)
commentLabel = Label(root, text=config['commentLabel'])
commentLabel.grid(row=2, column=0, pady=5)
textInputButton = Entry(root, bd=1)
textInputButton.grid(row=2, column=1, pady=5)
b8 = Button(root, text=config['writeToExcel'], command=writeToExcel, width=13)
b8.grid(row=3, column=0, columnspan=2)
b9 = Button(root, text=config['writeToExcelClose'], command=writeToExcelClose, width=13)
b9.grid(row=3, column=1, columnspan=2)

#Give focus to the text input area.
textInputButton.focus()

#This keeps the program running until the windows is crossed out by pressing the 'X' button.
root.mainloop()
